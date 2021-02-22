import openpyxl
import datetime

today = datetime.datetime.today()

# today +=  datetime.timedelta(days = 8)

month = today.strftime("%B")
weekday = int(today.strftime("%w"))
weekday += 6
weekday = weekday % 7

day = today.day

print(month, day, weekday)

wb = openpyxl.Workbook()
ws = wb.active

raw_data = openpyxl.load_workbook('./data/grafik.xlsx')
select_sheet = raw_data['Schedule']


# Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


for col in range(select_sheet.min_column, select_sheet.max_column + 1):
    if select_sheet.cell(3, col).value == month:
        start_column = col
        break


# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


print(select_sheet.cell(row=18, column=start_column + day - 1 + 7 - weekday).value)

range_to_copy = copyRange(startCol=+ day - 1 + 7 - weekday, startRow=18,
                          endCol=start_column + day - 1 + 7 - weekday + 7, endRow=18, sheet=select_sheet)
pasteRange(startCol=1, endCol=7, startRow=1, endRow=1, sheetReceiving=ws, copiedData=range_to_copy)
range_to_copy = copyRange(startCol=+ day - 1 + 7 - weekday, startRow=19,
                          endCol=start_column + day - 1 + 7 - weekday + 7, endRow=19, sheet=select_sheet)
pasteRange(startCol=8, endCol=14, startRow=1, endRow=1, sheetReceiving=ws, copiedData=range_to_copy)

range_to_copy = copyRange(startCol=1, startRow=2, endCol=3, endRow=2, sheet=select_sheet)
pasteRange(startCol=1, endCol=3, startRow=2, endRow=2, sheetReceiving=ws, copiedData=range_to_copy)

range_to_copy = copyRange(startCol=1, startRow=16, endCol=3, endRow=16, sheet=select_sheet)
pasteRange(startCol=1, endCol=3, startRow=3, endRow=3, sheetReceiving=ws, copiedData=range_to_copy)

people = 0
for row in range(select_sheet.min_row, select_sheet.max_row + 1):
    for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
        if select_sheet.cell(row, col).value == 'x':
            people += 1
            print(select_sheet.cell(row=row, column=1).value)
            ws.cell(column=1, row=3 + people).value = select_sheet.cell(row=row, column=1).value
            range_to_copy = copyRange(startCol=start_column + day - 1 + 7 - weekday, startRow=row,
                                      endCol=start_column + day - 1 + 7 - weekday + 7, endRow=row, sheet=select_sheet)
            pasteRange(startCol=8, endCol=14, startRow=3+people, endRow=3+people, sheetReceiving=ws, copiedData=range_to_copy)

wb.save('check.xlsx')
