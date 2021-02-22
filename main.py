
import ssl, smtplib, csv
from tabulate import tabulate
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import datetime


stmp_server= 'stmp.gmail.com'
port = 465
sender = input ("Enter your email")
password = input ("Enter your password")
message = MIMEMultipart("alternative")
message['Subject'] = '24/7 support - upcoming shift'

wb = openpyxl.load_workbook('./data/grafik.xlsx')


def send_email(to, msg):

    try:
        data.reverse()
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender, password)
        server.sendmail(sender, to, msg.as_string())
        server.quit()
    except Exception as e:
        print(e)
        print('Email Failed to Send to: ', )
        print("@: ", to)

def build_email(data, to_send):
    text = """
    Hi, 
    According to the agreement, You will be part of the 24/7 Support Team during net week.
    Please find the information about your team.

    {table}

    Best regards,

    Admin"""
    html = """
    <html>
    <head>
    <style> 
      table, th, td {{ border: 1px solid black; border-collapse: collapse; }}
      th, td {{ padding: 5px; }}
    </style>
    </head>
    <body><p>Hi, </p>
    <p>According to the agreement, You will be part of the 24/7 Support Team during net week.</p>
    <p>Please find the information about your team.</p>
    {table}
    <p>Best regards,</p>
    <p>Admin</p>
    </body></html>
    """
    #to = to_send.split(",")
    to = to_send
    text = text.format(table=tabulate(data, headers="firstrow", tablefmt="grid"))
    html = html.format(table=tabulate(data, headers="firstrow", tablefmt="html"))
    part1 = MIMEText(text, 'plain')
    part2 = MIMEText(html, 'html')
    message.attach(part1)
    message.attach(part2)
    send_email(to, message)


def get_data(wb):
    data = []
    select_sheet = wb['Schedule']
    merged = select_sheet.merged_cell_ranges
    for crange in merged:
        col_start, row_start, col_end, row_end = crange.bounds
        select_sheet.unmerge_cells(start_column=col_start, start_row=row_start, end_row=row_end, end_column=col_end)
        copied_value = select_sheet.cell(column=col_start, row=row_start).value
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                select_sheet.cell(column=col, row=row).value = copied_value

    today = datetime.datetime.today()
    month = today.strftime("%B")
    weekday = int(today.strftime("%w"))
    weekday += 6
    weekday = weekday % 7
    day = today.day
    workers = []
    for col in range(select_sheet.min_column, select_sheet.max_column + 1):
        if select_sheet.cell(3, col).value == month:
            start_column = col
            break
    for row in range(4, 15):
        data_to_append = []
        for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
            if select_sheet.cell(row, col).value == 'x':
                data_to_append.append(select_sheet.cell(row=4, column=col).value)
                data_to_append.append("")
                for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
                    data_to_append.append(select_sheet.cell(row=5, column=col).value)
                break
        if data_to_append:
            data.append(data_to_append)
    data.append(["MoC"])
    for row in range(4, 15):
        data_to_append = []
        for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
            if select_sheet.cell(row, col).value == 'x':
                data_to_append.append(select_sheet.cell(row=row, column=1).value)
                data_to_append.append("")
                for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
                    data_to_append.append(select_sheet.cell(row=row, column=col).value)
                workers.append(select_sheet.cell(row=row, column=2).value)
                break
        if data_to_append:
            data.append(data_to_append)
    data.append(["Tech"])
    for row in range(15, select_sheet.max_row+1):
        data_to_append = []
        for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
            if select_sheet.cell(row, col).value == 'x':
                workers.append(select_sheet.cell(row=row, column=2).value)
                data_to_append.append(select_sheet.cell(row=row, column=1).value)
                data_to_append.append("")
                for col in range(start_column + day - 1 + 7 - weekday, start_column + day - 1 + 7 - weekday + 7):
                    data_to_append.append(select_sheet.cell(row=row, column=col).value)
                break
        if data_to_append:
            data.append(data_to_append)
    return data, workers



data, workers = get_data(wb)
build_email(data, to_send=workers)
